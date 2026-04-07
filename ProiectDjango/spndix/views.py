from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from .models import (
    Cheltuiala,
    Categorie,
    Budget,
    AIAnaliza,
    ExportLog,
    ScanareLog,
    UserProfile,
    UserPlan,
    ForecastAlert,
    SavingsGoal,
    GoalContribution,
    Subscription,
    Household,
    HouseholdMember,
    OnboardingJourney,
    SmartAction,
    ReceiptInsight,
    LUNA_CHOICES,
)
from .forms import (
    CheltuialaForm,
    RegisterForm,
    BudgetForm,
    UserProfileForm,
    SavingsGoalForm,
    GoalContributionForm,
    SubscriptionForm,
    HouseholdCreateForm,
    HouseholdAddMemberForm,
)
from .plan_limits import check_limit, obtine_user_plan
from django.contrib import messages
from django.contrib.auth import login
from django.db.models import Sum, Q
from django.utils import timezone
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
import io
import json
import base64
import re
import uuid
from io import BytesIO
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from datetime import date, timedelta
from calendar import monthrange

import PIL.Image
import fitz
from PIL import ImageOps, UnidentifiedImageError
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
import google.generativeai as genai
import stripe


def construieste_buget_dashboard(buget, utilizator):
    suma_cheltuita = Cheltuiala.objects.filter(
        utilizator=utilizator,
        categorie=buget.categorie,
        data__month=buget.luna,
        data__year=buget.an,
    ).aggregate(total=Sum('suma'))['total'] or 0

    suma_limita = float(buget.suma_limita)
    suma_cheltuita_float = float(suma_cheltuita)
    procent = 0 if suma_limita == 0 else (suma_cheltuita_float / suma_limita) * 100
    progres = min(procent, 100)

    if procent > 100:
        clasa = 'bg-danger'
        status = 'Depășit'
    elif procent >= 80:
        clasa = 'bg-warning text-dark'
        status = 'Aproape de limită'
    else:
        clasa = 'bg-success'
        status = 'În regulă'

    return {
        'id': buget.pk,
        'categorie': buget.categorie,
        'luna': buget.luna,
        'luna_display': buget.get_luna_display(),
        'an': buget.an,
        'suma_limita': suma_limita,
        'suma_cheltuita': suma_cheltuita_float,
        'procent': round(procent, 1),
        'progres': round(progres, 1),
        'bar_class': clasa,
        'status': status,
        'excedat': procent > 100,
        'ramas': round(max(suma_limita - suma_cheltuita_float, 0), 2),
        'depasire': round(max(suma_cheltuita_float - suma_limita, 0), 2),
    }


def rotunjeste_bani(valoare):
    return Decimal(valoare or 0).quantize(Decimal('0.01'))


def curata_markdown(text):
    text = text or ''
    text = re.sub(r'#{1,6}\s+', '', text)
    text = re.sub(r'\*\*(.*?)\*\*', r'\1', text)
    text = re.sub(r'\*(.*?)\*', r'\1', text)
    text = re.sub(r'__(.*?)__', r'\1', text)
    return text.strip()


def suma_lunara_subscription(subscription):
    return rotunjeste_bani(subscription.suma_estimata)


def ziua_reala_subscription(subscription, an, luna):
    zi_setata = int(subscription.ziua_lunii or 1)
    zi_setata = max(1, zi_setata)
    zile_in_luna = monthrange(an, luna)[1]
    return min(zi_setata, zile_in_luna)


def urmatoarea_data_subscription(subscription, referinta=None):
    referinta = referinta or timezone.localdate()
    zi_curenta = ziua_reala_subscription(subscription, referinta.year, referinta.month)
    data_curenta = date(referinta.year, referinta.month, zi_curenta)
    if data_curenta >= referinta:
        return data_curenta

    if referinta.month == 12:
        an_urmator = referinta.year + 1
        luna_urmatoare = 1
    else:
        an_urmator = referinta.year
        luna_urmatoare = referinta.month + 1

    zi_urmatoare = ziua_reala_subscription(subscription, an_urmator, luna_urmatoare)
    return date(an_urmator, luna_urmatoare, zi_urmatoare)


def sincronizeaza_urmatoare_plati_subscriptions(utilizator, referinta=None):
    referinta = referinta or timezone.localdate()
    subscriptions = Subscription.objects.filter(utilizator=utilizator)

    for subscription in subscriptions:
        update_fields = []
        if subscription.ziua_lunii is None or int(subscription.ziua_lunii) < 1:
            subscription.ziua_lunii = 1
            update_fields.append('ziua_lunii')

        urmatoarea_plata = None
        if subscription.activ:
            urmatoarea_plata = urmatoarea_data_subscription(subscription, referinta=referinta)

        if subscription.urmatoarea_plata != urmatoarea_plata:
            subscription.urmatoarea_plata = urmatoarea_plata
            update_fields.append('urmatoarea_plata')

        if update_fields:
            subscription.save(update_fields=update_fields)


def auto_adauga_subscriptions_lunare(utilizator, azi=None):
    azi = azi or timezone.localdate()
    subscriptions = Subscription.objects.filter(utilizator=utilizator, activ=True).select_related('categorie')
    create_count = 0

    for subscription in subscriptions:
        ziua_reala = ziua_reala_subscription(subscription, azi.year, azi.month)
        data_scadenta = date(azi.year, azi.month, ziua_reala)

        if data_scadenta != azi:
            urmatoarea_plata = urmatoarea_data_subscription(subscription, referinta=azi)
            if subscription.urmatoarea_plata != urmatoarea_plata:
                subscription.urmatoarea_plata = urmatoarea_plata
                subscription.save(update_fields=['urmatoarea_plata'])
            continue

        exista_cheltuiala_luna_curenta = Cheltuiala.objects.filter(
            utilizator=utilizator,
            titlu__iexact=subscription.nume,
            data__year=azi.year,
            data__month=azi.month,
        ).exists()

        if not exista_cheltuiala_luna_curenta:
            Cheltuiala.objects.create(
                utilizator=utilizator,
                categorie=subscription.categorie,
                titlu=subscription.nume,
                suma=rotunjeste_bani(subscription.suma_estimata),
                data=azi,
                descriere='Abonament adaugat automat din Subscription Radar.',
            )
            create_count += 1

        subscription.ultima_plata = azi
        subscription.urmatoarea_plata = urmatoarea_data_subscription(
            subscription,
            referinta=azi + timedelta(days=1),
        )
        subscription.save(update_fields=['ultima_plata', 'urmatoarea_plata'])

    return create_count


def creeaza_alerta_daca_nu_exista(
    utilizator,
    categorie,
    tip,
    mesaj,
    suma_implicata,
    zile_ramase,
    actiune_recomandata,
):
    azi = timezone.localdate()
    exista = ForecastAlert.objects.filter(
        utilizator=utilizator,
        categorie=categorie,
        tip=tip,
        mesaj=mesaj,
        creat_la__date=azi,
    ).exists()
    if exista:
        return None

    return ForecastAlert.objects.create(
        utilizator=utilizator,
        categorie=categorie,
        tip=tip,
        mesaj=mesaj,
        suma_implicata=rotunjeste_bani(suma_implicata),
        zile_ramase=zile_ramase,
        actiune_recomandata=actiune_recomandata,
    )


def calculeaza_zile_pana_depasire(suma_limita, total_cheltuit, ritm_zilnic, zile_ramase_luna):
    if ritm_zilnic <= 0:
        return None

    suma_ramasa = Decimal(suma_limita or 0) - Decimal(total_cheltuit or 0)
    if suma_ramasa <= 0:
        return 0

    zile_float = float(suma_ramasa / ritm_zilnic)
    zile_int = int(zile_float)
    if zile_float > zile_int:
        zile_int += 1

    if zile_int > zile_ramase_luna:
        return None

    return max(zile_int, 0)


def luna_an_in_urma(luna, an, luni):
    luna_tinta = luna
    an_tinta = an
    pasi = max(int(luni or 0), 0)

    for _ in range(pasi):
        if luna_tinta == 1:
            luna_tinta = 12
            an_tinta -= 1
        else:
            luna_tinta -= 1

    return luna_tinta, an_tinta


KEYWORDS_SUBSCRIPTII_ESENTIALE = (
    'chirie', 'rent', 'rata', 'mortgage', 'utilitati', 'utility',
    'electricitate', 'electricity', 'gaz', 'gas', 'apa', 'water',
    'internet', 'telefon', 'phone', 'asigurare', 'insurance',
    'intretinere', 'condominiu', 'maintenance',
)


def subscription_este_esential(subscription):
    nume = (subscription.nume or '').strip().lower()
    categorie = ((subscription.categorie.nume if subscription.categorie else '') or '').strip().lower()
    text = f"{nume} {categorie}".strip()
    return any(keyword in text for keyword in KEYWORDS_SUBSCRIPTII_ESENTIALE)


def curata_sugestii_anulare_pentru_esentiale(utilizator, subscriptions):
    subscriptions_esentiale = [item for item in subscriptions if subscription_este_esential(item)]
    if not subscriptions_esentiale:
        return

    expresie = Q()
    for item in subscriptions_esentiale:
        expresie |= Q(mesaj__icontains=item.nume)
        if item.categorie and item.categorie.nume:
            expresie |= Q(mesaj__icontains=item.categorie.nume)

    if not expresie:
        return

    alerte_gresite = ForecastAlert.objects.filter(
        utilizator=utilizator,
        tip='sugestie_anulare',
        citita=False,
    ).filter(expresie)
    alerta_ids = list(alerte_gresite.values_list('id', flat=True))
    if not alerta_ids:
        return

    alerte_gresite.update(citita=True)
    SmartAction.objects.filter(
        utilizator=utilizator,
        alerta_id__in=alerta_ids,
        status='pending',
    ).update(status='dismissed')


def evalueaza_subscription_radar_avansat(utilizator, azi=None):
    azi = azi or timezone.localdate()
    subscriptions = list(
        Subscription.objects.filter(utilizator=utilizator, activ=True)
        .select_related('categorie')
        .order_by('-suma_estimata')
    )
    if not subscriptions:
        return

    curata_sugestii_anulare_pentru_esentiale(utilizator, subscriptions)

    for subscription in subscriptions:
        este_esential = subscription_este_esential(subscription)
        zile_pana_plata = None
        if subscription.urmatoarea_plata:
            zile_pana_plata = (subscription.urmatoarea_plata - azi).days

        if zile_pana_plata is not None and 0 <= zile_pana_plata <= 3:
            if este_esential:
                mesaj = (
                    f"Cheltuiala recurentă esențială {subscription.nume} este scadentă în {zile_pana_plata} "
                    f"zile (~{rotunjeste_bani(subscription.suma_estimata)} RON)."
                )
                actiune = (
                    f"Asigură fondurile pentru {subscription.nume} și, dacă poți, optimizează costul prin "
                    "renegociere sau verificarea ofertelor alternative."
                )
            else:
                mesaj = (
                    f"Abonamentul {subscription.nume} urmează să fie debitat în {zile_pana_plata} "
                    f"zile, cu aproximativ {rotunjeste_bani(subscription.suma_estimata)} RON."
                )
                actiune = (
                    f"Verifică dacă folosești {subscription.nume} și decide dacă păstrezi sau schimbi planul "
                    "înainte de următoarea debitare."
                )
            creeaza_alerta_daca_nu_exista(
                utilizator=utilizator,
                categorie=subscription.categorie,
                tip='abonament_iminent',
                mesaj=mesaj,
                suma_implicata=subscription.suma_estimata,
                zile_ramase=zile_pana_plata,
                actiune_recomandata=actiune,
            )

        cheltuiala_luna_curenta = Cheltuiala.objects.filter(
            utilizator=utilizator,
            titlu__iexact=subscription.nume,
            data__month=azi.month,
            data__year=azi.year,
        ).aggregate(total=Sum('suma'))['total'] or Decimal('0')

        totaluri_luni_anterioare = []
        for luni_anterioare in range(1, 4):
            luna_tinta, an_tinta = luna_an_in_urma(azi.month, azi.year, luni_anterioare)
            total_luna = Cheltuiala.objects.filter(
                utilizator=utilizator,
                titlu__iexact=subscription.nume,
                data__month=luna_tinta,
                data__year=an_tinta,
            ).aggregate(total=Sum('suma'))['total'] or Decimal('0')
            if total_luna > 0:
                totaluri_luni_anterioare.append(Decimal(total_luna))

        if totaluri_luni_anterioare and cheltuiala_luna_curenta > 0:
            medie_anterioara = sum(totaluri_luni_anterioare, Decimal('0')) / Decimal(len(totaluri_luni_anterioare))
            prag_scumpire = medie_anterioara * Decimal('1.15')
            if cheltuiala_luna_curenta >= prag_scumpire and medie_anterioara > 0:
                diferenta = rotunjeste_bani(cheltuiala_luna_curenta - medie_anterioara)
                crestere_pct = ((cheltuiala_luna_curenta - medie_anterioara) / medie_anterioara) * Decimal('100')
                crestere_text = crestere_pct.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)
                mesaj = (
                    f"Abonamentul {subscription.nume} pare mai scump luna aceasta: "
                    f"{rotunjeste_bani(cheltuiala_luna_curenta)} RON vs media recentă "
                    f"{rotunjeste_bani(medie_anterioara)} RON (+{crestere_text}%)."
                )
                if este_esential:
                    actiune = (
                        f"{subscription.nume} este o cheltuială esențială: prioritar renegociază contractul "
                        f"sau optimizează consumul. Poți recupera aproximativ {diferenta} RON/lună."
                    )
                else:
                    actiune = (
                        f"Verifică planul {subscription.nume}, caută opțiuni mai ieftine sau negociază prețul. "
                        f"Poți recupera aproximativ {diferenta} RON/lună."
                    )
                creeaza_alerta_daca_nu_exista(
                    utilizator=utilizator,
                    categorie=subscription.categorie,
                    tip='abonament_scumpit',
                    mesaj=mesaj,
                    suma_implicata=diferenta,
                    zile_ramase=zile_pana_plata,
                    actiune_recomandata=actiune,
                )

    try:
        profil = utilizator.userprofile
        venit_lunar = Decimal(profil.venit_lunar or 0)
    except UserProfile.DoesNotExist:
        venit_lunar = Decimal('0')

    total_subscriptions = sum((suma_lunara_subscription(item) for item in subscriptions), Decimal('0'))
    if venit_lunar > 0 and subscriptions and total_subscriptions >= venit_lunar * Decimal('0.25'):
        subscriptions_optionale = [item for item in subscriptions if not subscription_este_esential(item)]
        subscription_scump = subscriptions_optionale[0] if subscriptions_optionale else None

        if subscription_scump is None:
            mesaj = (
                f"Cheltuielile recurente active consumă {rotunjeste_bani(total_subscriptions)} RON/lună (peste 25% din venit), "
                "dar majoritatea sunt esențiale."
            )
            actiune = (
                "Nu anula cheltuielile esențiale (ex. chirie/utilități). Concentrează-te pe renegociere și pe reducerea "
                "abonamentelor opționale, dacă apar."
            )
            creeaza_alerta_daca_nu_exista(
                utilizator=utilizator,
                categorie=None,
                tip='ritm_alert',
                mesaj=mesaj,
                suma_implicata=total_subscriptions,
                zile_ramase=7,
                actiune_recomandata=actiune,
            )
            return

        zile_pana_plata = None
        if subscription_scump.urmatoarea_plata:
            zile_pana_plata = (subscription_scump.urmatoarea_plata - azi).days
        mesaj = (
            f"Abonamentele active consumă {rotunjeste_bani(total_subscriptions)} RON/lună, peste 25% din venitul tău. "
            f"Cel mai scump este {subscription_scump.nume} ({rotunjeste_bani(subscription_scump.suma_estimata)} RON)."
        )
        actiune = (
            f"Dacă anulezi sau reduci planul pentru {subscription_scump.nume}, poți elibera rapid "
            f"{rotunjeste_bani(subscription_scump.suma_estimata)} RON/lună."
        )
        creeaza_alerta_daca_nu_exista(
            utilizator=utilizator,
            categorie=subscription_scump.categorie,
            tip='sugestie_anulare',
            mesaj=mesaj,
            suma_implicata=subscription_scump.suma_estimata,
            zile_ramase=zile_pana_plata,
            actiune_recomandata=actiune,
        )


def sincronizeaza_smart_actions_din_alerte(utilizator):
    azi = timezone.localdate()

    alerte_active = ForecastAlert.objects.filter(
        utilizator=utilizator,
        citita=False,
    ).select_related('categorie').order_by('-creat_la')[:30]

    for alerta in alerte_active:
        if alerta.tip in ['depasire_iminenta', 'ritm_alert']:
            categorie_text = alerta.categorie.nume if alerta.categorie else 'categoria relevantă'
            titlu = f"Optimizează bugetul pentru {categorie_text}"
            tip_action = 'buget'
        elif alerta.tip == 'abonament_iminent':
            categorie_text = alerta.categorie.nume if alerta.categorie else 'abonamente'
            titlu = f"Pregătește plata recurentă ({categorie_text})"
            tip_action = 'abonament'
        elif alerta.tip == 'abonament_scumpit':
            titlu = 'Verifică scumpirea unui abonament'
            tip_action = 'abonament'
        elif alerta.tip == 'sugestie_anulare':
            titlu = 'Evaluează reducerea abonamentelor opționale'
            tip_action = 'abonament'
        elif alerta.tip == 'recurenta_detectata':
            titlu = 'Confirmă o cheltuială recurentă detectată'
            tip_action = 'abonament'
        else:
            titlu = 'Aplică o optimizare financiară recomandată'
            tip_action = 'economii'

        data_scadenta = None
        if alerta.zile_ramase is not None:
            data_scadenta = azi + timedelta(days=max(1, min(int(alerta.zile_ramase), 14)))
        else:
            data_scadenta = azi + timedelta(days=3)

        defaults = {
            'utilizator': utilizator,
            'titlu': titlu,
            'descriere': alerta.actiune_recomandata,
            'tip': tip_action,
            'status': 'pending',
            'impact_estimat': rotunjeste_bani(alerta.suma_implicata),
            'data_scadenta': data_scadenta,
        }

        action, created = SmartAction.objects.get_or_create(
            alerta=alerta,
            defaults=defaults,
        )
        if not created and action.status == 'pending':
            action.titlu = titlu
            action.descriere = alerta.actiune_recomandata
            action.tip = tip_action
            action.impact_estimat = rotunjeste_bani(alerta.suma_implicata)
            action.data_scadenta = data_scadenta
            action.save(update_fields=['titlu', 'descriere', 'tip', 'impact_estimat', 'data_scadenta'])

    SmartAction.objects.filter(
        utilizator=utilizator,
        alerta__citita=True,
        status='pending',
    ).update(status='dismissed')

    return list(
        SmartAction.objects.filter(utilizator=utilizator)
        .select_related('alerta')
        .order_by('status', 'data_scadenta', '-creat_la')[:12]
    )


def media_decimala(valori):
    valori_filtrate = [Decimal(val) for val in valori if Decimal(val) > 0]
    if not valori_filtrate:
        return Decimal('0')
    return sum(valori_filtrate, Decimal('0')) / Decimal(len(valori_filtrate))


def calculeaza_receipt_intelligence(utilizator, azi=None):
    azi = azi or timezone.localdate()
    start = azi - timedelta(days=120)
    bonuri = list(
        ReceiptInsight.objects.filter(utilizator=utilizator, data_bon__gte=start)
        .order_by('-data_bon', '-creat_la')
    )
    if not bonuri:
        return None

    costuri_pe_magazin = {}
    for bon in bonuri:
        cheia = (bon.magazin or 'Necunoscut').strip() or 'Necunoscut'
        cost_unitar = Decimal(bon.pret_mediu_produs or 0)
        if cost_unitar <= 0:
            produse = max(int(bon.nr_produse or 0), 1)
            cost_unitar = Decimal(bon.total or 0) / Decimal(produse)
        costuri_pe_magazin.setdefault(cheia, []).append(cost_unitar)

    magazine_medii = {
        magazin: media_decimala(valori)
        for magazin, valori in costuri_pe_magazin.items()
        if valori
    }
    magazin_best = min(magazine_medii, key=magazine_medii.get)
    media_best = magazine_medii[magazin_best]
    media_globala = media_decimala(list(magazine_medii.values()))

    economisire_pct = Decimal('0')
    if media_globala > 0:
        economisire_pct = ((media_globala - media_best) / media_globala) * Decimal('100')

    perioada_curenta_start = azi - timedelta(days=30)
    perioada_anterioara_start = azi - timedelta(days=60)
    perioada_anterioara_end = azi - timedelta(days=31)

    preturi_curente = [
        Decimal(item.pret_mediu_produs or 0)
        for item in bonuri
        if item.data_bon >= perioada_curenta_start and Decimal(item.pret_mediu_produs or 0) > 0
    ]
    preturi_anterioare = [
        Decimal(item.pret_mediu_produs or 0)
        for item in bonuri
        if perioada_anterioara_start <= item.data_bon <= perioada_anterioara_end and Decimal(item.pret_mediu_produs or 0) > 0
    ]
    medie_curenta = media_decimala(preturi_curente)
    medie_anterioara = media_decimala(preturi_anterioare)

    inflatie_personala_pct = None
    if medie_anterioara > 0 and medie_curenta > 0:
        inflatie_personala_pct = ((medie_curenta - medie_anterioara) / medie_anterioara) * Decimal('100')

    luna_precedenta, an_precedent = obtine_luna_precedenta(azi.month, azi.year)
    cos_curent = media_decimala([
        Decimal(item.total or 0)
        for item in bonuri
        if item.data_bon.month == azi.month and item.data_bon.year == azi.year
    ])
    cos_precedent = media_decimala([
        Decimal(item.total or 0)
        for item in bonuri
        if item.data_bon.month == luna_precedenta and item.data_bon.year == an_precedent
    ])

    variatie_cos_pct = None
    if cos_precedent > 0 and cos_curent > 0:
        variatie_cos_pct = ((cos_curent - cos_precedent) / cos_precedent) * Decimal('100')

    return {
        'bonuri_count': len(bonuri),
        'magazin_best': magazin_best,
        'cost_mediu_best': rotunjeste_bani(media_best),
        'economisire_pct': economisire_pct.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP),
        'inflatie_personala_pct': None if inflatie_personala_pct is None else inflatie_personala_pct.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP),
        'cos_mediu_curent': rotunjeste_bani(cos_curent),
        'cos_mediu_precedent': rotunjeste_bani(cos_precedent),
        'variatie_cos_pct': None if variatie_cos_pct is None else variatie_cos_pct.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP),
    }


def obtine_household_summary(utilizator):
    membership = HouseholdMember.objects.filter(
        utilizator=utilizator,
        activ=True,
    ).select_related('household').first()
    if not membership:
        return None

    household = membership.household
    membri = list(household.membri.filter(activ=True).select_related('utilizator'))
    member_ids = [item.utilizator_id for item in membri]

    azi = timezone.localdate()
    total_luna = Cheltuiala.objects.filter(
        utilizator_id__in=member_ids,
        data__month=azi.month,
        data__year=azi.year,
    ).aggregate(total=Sum('suma'))['total'] or Decimal('0')

    return {
        'household': household,
        'rol_curent': membership.rol,
        'membri_count': len(membri),
        'total_luna': rotunjeste_bani(total_luna),
    }


def detecteaza_recurente(utilizator, azi):
    inceput_istoric = azi - timedelta(days=210)
    cheltuieli_istorice = Cheltuiala.objects.filter(
        utilizator=utilizator,
        data__gte=inceput_istoric,
        data__lt=azi,
    ).select_related('categorie').order_by('titlu', 'data')

    grupuri = {}
    for cheltuiala in cheltuieli_istorice:
        titlu_normalizat = (cheltuiala.titlu or '').strip().lower()
        if not titlu_normalizat:
            continue
        grupuri.setdefault(titlu_normalizat, []).append(cheltuiala)

    for _, elemente in grupuri.items():
        luni_distincte = {(item.data.year, item.data.month) for item in elemente}
        if len(luni_distincte) < 2:
            continue

        zile = [item.data.day for item in elemente]
        if max(zile) - min(zile) > 7:
            continue

        zi_medie = max(1, round(sum(zile) / len(zile)))
        exista_luna_curenta = any(item.data.year == azi.year and item.data.month == azi.month for item in elemente)

        if not exista_luna_curenta and azi.day <= zi_medie:
            an_tinta = azi.year
            luna_tinta = azi.month
        else:
            if azi.month == 12:
                an_tinta = azi.year + 1
                luna_tinta = 1
            else:
                an_tinta = azi.year
                luna_tinta = azi.month + 1

        zi_tinta = min(zi_medie, monthrange(an_tinta, luna_tinta)[1])
        data_urmatoare = date(an_tinta, luna_tinta, zi_tinta)
        zile_ramase = (data_urmatoare - azi).days

        if zile_ramase < 0 or zile_ramase > 40:
            continue

        ultima_cheltuiala = elemente[-1]
        suma_estimata = rotunjeste_bani(sum(item.suma for item in elemente) / Decimal(len(elemente)))
        mesaj = (
            f"Cheltuială recurentă detectată: {ultima_cheltuiala.titlu} apare în fiecare lună. "
            f"Urmează să apară în aproximativ {zile_ramase} zile cu o sumă estimată de {suma_estimata} RON."
        )
        actiune = 'Marchează ca așteptat sau setează reminder.'
        creeaza_alerta_daca_nu_exista(
            utilizator=utilizator,
            categorie=ultima_cheltuiala.categorie,
            tip='recurenta_detectata',
            mesaj=mesaj,
            suma_implicata=suma_estimata,
            zile_ramase=zile_ramase,
            actiune_recomandata=actiune,
        )


def calculate_forecasts(utilizator):
    azi = timezone.localdate()
    luna_curenta = azi.month
    an_curent = azi.year
    zile_trecute = max(azi.day, 1)
    zile_in_luna = monthrange(an_curent, luna_curenta)[1]
    zile_ramase_luna = max(zile_in_luna - zile_trecute, 0)
    data_sf_luna = date(an_curent, luna_curenta, zile_in_luna)

    totaluri_categorii = Cheltuiala.objects.filter(
        utilizator=utilizator,
        data__month=luna_curenta,
        data__year=an_curent,
    ).values('categorie').annotate(total=Sum('suma'))

    totaluri_map = {
        item['categorie']: Decimal(item['total'] or 0)
        for item in totaluri_categorii
    }

    luna_precedenta, an_precedent = (12, an_curent - 1) if luna_curenta == 1 else (luna_curenta - 1, an_curent)
    totaluri_precedente = Cheltuiala.objects.filter(
        utilizator=utilizator,
        data__month=luna_precedenta,
        data__year=an_precedent,
    ).values('categorie').annotate(total=Sum('suma'))

    totaluri_precedente_map = {
        item['categorie']: Decimal(item['total'] or 0)
        for item in totaluri_precedente
    }

    bugete_curente = Budget.objects.filter(
        utilizator=utilizator,
        luna=luna_curenta,
        an=an_curent,
    ).select_related('categorie')

    for buget in bugete_curente:
        suma_limita = Decimal(buget.suma_limita or 0)
        if suma_limita <= 0:
            continue

        total_cheltuit = Decimal(totaluri_map.get(buget.categorie_id, Decimal('0')) or 0)
        ritm_zilnic = total_cheltuit / Decimal(zile_trecute)
        proiectie_luna = ritm_zilnic * Decimal(zile_in_luna)
        procent_folosit = (total_cheltuit / suma_limita) * Decimal('100')
        procent_text = procent_folosit.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)
        suma_ramasa = rotunjeste_bani(max(suma_limita - total_cheltuit, Decimal('0')))
        zile_pana_depasire = calculeaza_zile_pana_depasire(
            suma_limita=suma_limita,
            total_cheltuit=total_cheltuit,
            ritm_zilnic=ritm_zilnic,
            zile_ramase_luna=zile_ramase_luna,
        )

        suma_depasire = rotunjeste_bani(max(proiectie_luna - suma_limita, Decimal('0')))
        if suma_depasire > 0:
            if zile_pana_depasire is not None:
                mesaj = (
                    f"La ritmul actual, vei depăși bugetul pentru {buget.categorie.nume} în aproximativ "
                    f"{zile_pana_depasire} zile. Depășirea estimată până la finalul lunii este de "
                    f"{suma_depasire} RON."
                )
            else:
                mesaj = (
                    f"La ritmul actual, vei depăși bugetul de {rotunjeste_bani(suma_limita)} RON pentru "
                    f"{buget.categorie.nume} cu aproximativ {suma_depasire} RON până la sfârșitul lunii."
                )
            actiune = (
                f"Intră pe cheltuielile din {buget.categorie.nume} și prioritizează doar costurile "
                "esențiale pentru restul lunii."
            )
            creeaza_alerta_daca_nu_exista(
                utilizator=utilizator,
                categorie=buget.categorie,
                tip='depasire_iminenta',
                mesaj=mesaj,
                suma_implicata=suma_depasire,
                zile_ramase=zile_pana_depasire if zile_pana_depasire is not None else zile_ramase_luna,
                actiune_recomandata=actiune,
            )
            continue

        if proiectie_luna > suma_limita * Decimal('0.85') and proiectie_luna < suma_limita:
            mesaj = (
                f"Ai cheltuit {procent_text}% din bugetul pentru {buget.categorie.nume} și mai sunt "
                f"{zile_ramase_luna} zile din lună. Încearcă să limitezi cheltuielile la {suma_ramasa} RON "
                f"până pe {data_sf_luna.strftime('%d.%m.%Y')}."
            )
            actiune = (
                f"Verifică zilnic categoria {buget.categorie.nume} și păstrează cheltuielile în limita de "
                f"{suma_ramasa} RON pentru restul lunii."
            )
            creeaza_alerta_daca_nu_exista(
                utilizator=utilizator,
                categorie=buget.categorie,
                tip='ritm_alert',
                mesaj=mesaj,
                suma_implicata=suma_ramasa,
                zile_ramase=zile_ramase_luna,
                actiune_recomandata=actiune,
            )

        if total_cheltuit < suma_limita * Decimal('0.5') and zile_trecute > 15:
            total_luna_precedenta = Decimal(totaluri_precedente_map.get(buget.categorie_id, Decimal('0')) or 0)
            economie = rotunjeste_bani(max(total_luna_precedenta - proiectie_luna, Decimal('0')))
            if economie > 0:
                mesaj = (
                    f"Ai cheltuit doar {procent_text}% din bugetul pentru {buget.categorie.nume}. "
                    f"Poți economisi până la {economie} RON față de luna trecută dacă menții același ritm."
                )
                actiune = (
                    f"Menține ritmul curent în {buget.categorie.nume}; proiecția lunară este "
                    f"{rotunjeste_bani(proiectie_luna)} RON."
                )
                creeaza_alerta_daca_nu_exista(
                    utilizator=utilizator,
                    categorie=buget.categorie,
                    tip='economie_posibila',
                    mesaj=mesaj,
                    suma_implicata=economie,
                    zile_ramase=zile_ramase_luna,
                    actiune_recomandata=actiune,
                )

    detecteaza_recurente(utilizator, azi)
    evalueaza_subscription_radar_avansat(utilizator, azi=azi)

    return list(
        ForecastAlert.objects.filter(utilizator=utilizator, citita=False)
        .select_related('categorie')
        .order_by('-creat_la')[:8]
    )


def procent_goal(suma_curenta, suma_tinta):
    suma_tinta = Decimal(suma_tinta or 0)
    if suma_tinta <= 0:
        return Decimal('0')
    return (Decimal(suma_curenta or 0) / suma_tinta) * Decimal('100')


def construieste_goal_status(goal, contributii_saptamana_map=None):
    azi = timezone.localdate()
    suma_tinta = Decimal(goal.suma_tinta or 0)
    suma_curenta = Decimal(goal.suma_curenta or 0)
    procent_real = procent_goal(suma_curenta, suma_tinta)
    procent_limitat = min(procent_real, Decimal('100'))
    suma_ramasa = rotunjeste_bani(max(suma_tinta - suma_curenta, Decimal('0')))
    contributie_saptamana = Decimal('0')
    if contributii_saptamana_map is not None:
        contributie_saptamana = Decimal(contributii_saptamana_map.get(goal.pk, Decimal('0')) or 0)

    zile_ramase = None
    if goal.data_tinta:
        zile_ramase = (goal.data_tinta - azi).days

    milestones = []
    for prag in (25, 50, 75, 100):
        milestones.append({'prag': prag, 'atins': procent_real >= Decimal(prag)})

    return {
        'goal': goal,
        'suma_tinta': suma_tinta,
        'suma_curenta': suma_curenta,
        'suma_ramasa': suma_ramasa,
        'procent_real': float(procent_real),
        'procent': float(procent_limitat.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)),
        'zile_ramase': zile_ramase,
        'contributie_saptamana': rotunjeste_bani(contributie_saptamana),
        'milestones': milestones,
        'completat': suma_tinta > 0 and suma_curenta >= suma_tinta,
    }


def notifica_milestone_goal(goal, procent_inainte, procent_dupa):
    azi = timezone.localdate()
    zile_ramase = (goal.data_tinta - azi).days if goal.data_tinta else None
    suma_ramasa = rotunjeste_bani(max(Decimal(goal.suma_tinta or 0) - Decimal(goal.suma_curenta or 0), Decimal('0')))

    for prag in (25, 50, 75, 100):
        if procent_inainte < prag <= procent_dupa:
            mesaj = f"Ai atins {prag}% din goal-ul '{goal.titlu}'! Mai ai nevoie de {suma_ramasa} RON."
            actiune = f"Continuă contribuțiile pentru '{goal.titlu}' până atingi ținta completă."
            creeaza_alerta_daca_nu_exista(
                utilizator=goal.utilizator,
                categorie=None,
                tip='economie_posibila',
                mesaj=mesaj,
                suma_implicata=suma_ramasa,
                zile_ramase=zile_ramase,
                actiune_recomandata=actiune,
            )


def obtine_onboarding_context(utilizator):
    journey, _ = OnboardingJourney.objects.get_or_create(utilizator=utilizator)

    try:
        profil = utilizator.userprofile
        profil_complet = bool(
            profil.tip_gospodarie and
            profil.obiectiv and
            profil.nr_persoane >= 1 and
            profil.venit_lunar is not None
        )
    except UserProfile.DoesNotExist:
        profil_complet = False

    pasi = [
        {
            'key': 'profil',
            'titlu': 'Completează profilul financiar',
            'descriere': 'Adaugă venit, tipul gospodăriei și obiectivul financiar.',
            'done': profil_complet,
            'url_name': 'profil',
        },
        {
            'key': 'buget',
            'titlu': 'Setează primul buget',
            'descriere': 'Definirea unui buget activează alertele predictive.',
            'done': Budget.objects.filter(utilizator=utilizator).exists(),
            'url_name': 'adauga_buget',
        },
        {
            'key': 'scan',
            'titlu': 'Scanează primul bon',
            'descriere': 'OCR-ul îți alimentează automat istoricul de cheltuieli.',
            'done': ScanareLog.objects.filter(utilizator=utilizator).exists(),
            'url_name': 'scaneaza_bon',
        },
        {
            'key': 'goal',
            'titlu': 'Creează un savings goal',
            'descriere': 'Setează o țintă ca să vezi progresul săptămânal.',
            'done': SavingsGoal.objects.filter(utilizator=utilizator).exists(),
            'url_name': 'adauga_goal',
        },
        {
            'key': 'subscription',
            'titlu': 'Adaugă un abonament recurent',
            'descriere': 'Subscription Radar poate preveni scurgerile de bani lunare.',
            'done': Subscription.objects.filter(utilizator=utilizator).exists(),
            'url_name': 'lista_subscriptions',
        },
    ]

    total_pasi = len(pasi)
    pasi_completati = sum(1 for pas in pasi if pas['done'])
    progres_pct = int((pasi_completati / total_pasi) * 100) if total_pasi else 0

    azi = timezone.localdate()
    if journey.data_tinta and journey.data_tinta < journey.data_start:
        journey.data_tinta = journey.data_start + timedelta(days=7)
        journey.save(update_fields=['data_tinta'])

    zile_ramase = None
    if journey.data_tinta:
        zile_ramase = (journey.data_tinta - azi).days

    if pasi_completati >= 3 and not journey.first_win_obtinut:
        journey.first_win_obtinut = True
        journey.first_win_la = timezone.now()
        journey.save(update_fields=['first_win_obtinut', 'first_win_la'])

    pending_onboarding_action = SmartAction.objects.filter(
        utilizator=utilizator,
        tip='onboarding',
        status='pending',
    ).order_by('creat_la').first()

    if not journey.first_win_obtinut and not journey.ascuns:
        urmatorul_pas = next((pas for pas in pasi if not pas['done']), None)
        if urmatorul_pas:
            titlu = f"Onboarding: {urmatorul_pas['titlu']}"
            descriere = urmatorul_pas['descriere']
            if pending_onboarding_action:
                if pending_onboarding_action.titlu != titlu or pending_onboarding_action.descriere != descriere:
                    pending_onboarding_action.titlu = titlu
                    pending_onboarding_action.descriere = descriere
                    pending_onboarding_action.data_scadenta = journey.data_tinta
                    pending_onboarding_action.save(update_fields=['titlu', 'descriere', 'data_scadenta'])
            else:
                SmartAction.objects.create(
                    utilizator=utilizator,
                    titlu=titlu,
                    descriere=descriere,
                    tip='onboarding',
                    status='pending',
                    impact_estimat=Decimal('0'),
                    data_scadenta=journey.data_tinta,
                )
    else:
        SmartAction.objects.filter(
            utilizator=utilizator,
            tip='onboarding',
            status='pending',
        ).update(status='done', completata_la=timezone.now())

    return {
        'journey': journey,
        'pasi': pasi,
        'pasi_completati': pasi_completati,
        'total_pasi': total_pasi,
        'progres_pct': progres_pct,
        'zile_ramase': zile_ramase,
        'afiseaza_card': (not journey.ascuns and not journey.first_win_obtinut),
    }


def obtine_filtre_luna_an(request):
    acum = timezone.now()
    luna = request.GET.get('month')
    an = request.GET.get('year')

    try:
        luna = int(luna)
    except (TypeError, ValueError):
        luna = acum.month

    try:
        an = int(an)
    except (TypeError, ValueError):
        an = acum.year

    if luna < 1 or luna > 12:
        luna = acum.month

    return luna, an


def cheltuieli_din_filtre(request):
    luna, an = obtine_filtre_luna_an(request)
    cheltuieli = Cheltuiala.objects.filter(
        utilizator=request.user,
        data__month=luna,
        data__year=an,
    ).select_related('categorie').order_by('-data', '-id')
    return cheltuieli, luna, an


def nume_fisier_export(luna, an, extensie):
    return f"spndix_cheltuieli_{luna:02d}_{an}.{extensie}"


def luna_display(luna):
    return dict(LUNA_CHOICES).get(luna, str(luna))


def luna_ani_disponibili(request):
    date_cheltuieli = Cheltuiala.objects.filter(utilizator=request.user).dates('data', 'year')
    ani = sorted({data.year for data in date_cheltuieli}, reverse=True)
    if not ani:
        ani = [timezone.now().year]
    return ani


def rezumat_cheltuieli_ai(request, luna, an):
    cheltuieli = Cheltuiala.objects.filter(
        utilizator=request.user,
        data__month=luna,
        data__year=an,
    ).select_related('categorie').order_by('-suma', 'categorie__nume')

    totaluri_categorii = list(
        cheltuieli.values('categorie__nume', 'categorie__culoare')
        .annotate(total=Sum('suma'))
        .order_by('-total')
    )

    bugete = Budget.objects.filter(
        utilizator=request.user,
        luna=luna,
        an=an,
    ).select_related('categorie')

    cheltuieli_detaliate = []
    for item in cheltuieli:
        cheltuieli_detaliate.append({
            'titlu': item.titlu,
            'categorie': item.categorie.nume if item.categorie else 'Fără categorie',
            'suma': float(item.suma),
            'data': item.data.strftime('%d.%m.%Y'),
            'descriere': item.descriere or '',
        })

    bugete_detaliate = []
    for buget in bugete:
        cheltuit = cheltuieli.filter(categorie=buget.categorie).aggregate(total=Sum('suma'))['total'] or Decimal('0')
        bugete_detaliate.append({
            'categorie': buget.categorie.nume,
            'suma_limita': float(buget.suma_limita),
            'suma_cheltuita': float(cheltuit),
            'procent': 0 if buget.suma_limita == 0 else round((float(cheltuit) / float(buget.suma_limita)) * 100, 1),
        })

    date_cheltuieli_text = json.dumps({
        'cheltuieli': cheltuieli_detaliate,
        'totaluri_pe_categorii': [
            {
                'categorie': item['categorie__nume'] or 'Fără categorie',
                'culoare': item['categorie__culoare'] or '#6c757d',
                'total': float(item['total']),
            }
            for item in totaluri_categorii
        ],
    }, ensure_ascii=False, indent=2)

    date_bugete_text = json.dumps(bugete_detaliate, ensure_ascii=False, indent=2)

    return date_cheltuieli_text, date_bugete_text, cheltuieli, totaluri_categorii, bugete_detaliate


def construieste_cheltuieli_detaliate(cheltuieli):
    linii = []
    for item in cheltuieli:
        categorie = item.categorie.nume if item.categorie else 'Fără categorie'
        linii.append(
            f"- [{item.data.strftime('%d.%m.%Y')}] {item.titlu} ({categorie}): {float(item.suma):.2f} RON"
        )
    return "\n".join(linii) if linii else "Nu există cheltuieli înregistrate pentru perioada selectată."


def obtine_luna_precedenta(luna, an):
    if luna == 1:
        return 12, an - 1
    return luna - 1, an


def construieste_date_luna_precedenta(utilizator, luna, an):
    luna_precedenta, an_precedent = obtine_luna_precedenta(luna, an)
    cheltuieli_precedente = Cheltuiala.objects.filter(
        utilizator=utilizator,
        data__month=luna_precedenta,
        data__year=an_precedent,
    ).select_related('categorie').order_by('-data', '-id')

    if not cheltuieli_precedente.exists():
        return None

    cheltuieli_text = construieste_cheltuieli_detaliate(cheltuieli_precedente)
    totaluri_categorii = list(
        cheltuieli_precedente.values('categorie__nume')
        .annotate(total=Sum('suma'))
        .order_by('-total')
    )

    if totaluri_categorii:
        totaluri_text = "\n".join(
            f"- {(item['categorie__nume'] or 'Fără categorie')}: {float(item['total']):.2f} RON"
            for item in totaluri_categorii
        )
    else:
        totaluri_text = "- Fără date"

    return (
        f"Luna precedentă analizată: {luna_display(luna_precedenta)} {an_precedent}\n"
        f"Cheltuieli detaliate:\n{cheltuieli_text}\n\n"
        "Totaluri pe categorii:\n"
        f"{totaluri_text}"
    )


def construieste_prompt_analiza(
    luna,
    an,
    tip_gospodarie,
    nr_persoane,
    are_copii,
    venit_lunar,
    obiectiv,
    date_cheltuieli_detaliate,
    date_bugete,
    date_luna_precedenta,
):
    prompt = f"""
Răspunde în text simplu fără Markdown, fără simboluri ### ** * _ sau alte formatări speciale.
Folosește doar text simplu cu paragrafe separate.

Ești un consilier financiar personal experimentat,
nu un chatbot generic. Analizezi cheltuielile reale
ale unui utilizator și dai sfaturi SPECIFICE bazate
EXCLUSIV pe datele lui, nu sfaturi generale.

CONTEXT UTILIZATOR:
- Tip gospodărie: {tip_gospodarie}
- Număr persoane în gospodărie: {nr_persoane}
- Are copii: {are_copii}
- Venit lunar aproximativ: {venit_lunar} RON
- Obiectiv financiar: {obiectiv}

CHELTUIELI LUNA {luna} {an}:
{date_cheltuieli_detaliate}

BUGETE SETATE:
{date_bugete}
"""

    if date_luna_precedenta:
        prompt += f"""
COMPARAȚIE CU LUNA PRECEDENTĂ:
{date_luna_precedenta}
Compară cheltuielile cu luna precedentă și identifică ce a crescut/scăzut.
"""
    else:
        prompt += """
NOTĂ: Nu există date din luna precedentă.
Nu face comparații cu luna trecută.
Analizează doar datele din luna curentă.
"""

    prompt += f"""

REGULI STRICTE:
1. NU da sfaturi generice - doar sfaturi bazate
   pe datele CONCRETE ale acestui utilizator
2. Calculează cheltuielile per persoană din
   gospodărie unde e relevant
3. Compară cu luna precedentă și identifică
   ce a crescut/scăzut
4. Identifică anomalii specifice cu sume exacte
5. Sfaturile trebuie să fie ACȚIONABILE cu sume exacte:
   NU: "Reduce cheltuielile pe divertisment"
   DA: "Ai plătit Netflix (45 RON) + HBO (35 RON) =
       80 RON. Dacă împarți Netflix cu cineva
       economisești 45 RON/lună = 540 RON/an"
6. Dacă sunt sub 10 cheltuieli introduse, menționează
   că analiza e limitată și îndeamnă să adauge mai multe
7. Ține cont de contextul familial - nu sugera reducerea
   cheltuielilor esențiale pentru familie dacă sunt
   rezonabile per persoană
8. Tonul să fie ca al unui prieten care ajută,
   nu ca un profesor care ceartă

STRUCTURĂ RĂSPUNS (respectă exact această structură):
1. Sumar luna {luna} {an} (3-4 rânduri cu totaluri exacte)
2. Top 3 categorii cu cele mai mari cheltuieli
   (sume exacte + comparație cu luna precedentă)
3. O anomalie sau pattern interesant observat în date
4. 3 sfaturi SPECIFICE și ACȚIONABILE cu sume exacte
5. Estimare economii: dacă urmezi sfaturile,
   cât poți economisi luna viitoare?

Răspunde în română, prietenos și direct.
"""

    return prompt


def genereaza_text_gemini(prompt):
    modele_candidate = (
        'gemini-2.0-flash',
        'gemini-2.0-flash-lite',
        'gemini-flash-lite-latest',
        'gemini-1.5-flash-latest',
    )
    ultima_eroare = None

    for model_name in modele_candidate:
        try:
            model = genai.GenerativeModel(model_name)
            response = model.generate_content(prompt)
            raspuns = (response.text or '').strip()
            if raspuns:
                return raspuns
            ultima_eroare = RuntimeError(f'Modelul {model_name} a returnat răspuns gol.')
        except Exception as exc:
            ultima_eroare = exc

    raise RuntimeError(
        'Nu am găsit un model Gemini compatibil pentru analiză. '
        f'Ultima eroare: {ultima_eroare}'
    )


def construieste_descriere_bon(produse, magazin, data_cumpararii, categorie_sugerata):
    linii = [f"{produs.get('nume', 'Produs')}: {float(produs.get('pret', 0)):.2f} RON" for produs in produse]
    bloc_produse = "\n".join(f"- {linie}" for linie in linii) if linii else "- Nu au fost extrase produse individuale."
    return (
        f"Bon scanat automat.\n"
        f"Magazin: {magazin}\n"
        f"Data extrasă: {data_cumpararii}\n"
        f"Categorie sugerată: {categorie_sugerata or '—'}\n\n"
        f"Produse extrase:\n{bloc_produse}"
    )


def curata_json_text(text):
    curat = text.strip()
    if curat.startswith('```'):
        curat = re.sub(r'^```(?:json)?\s*', '', curat, flags=re.IGNORECASE | re.DOTALL)
        curat = re.sub(r'\s*```$', '', curat, flags=re.DOTALL).strip()
    if not curat.startswith('{'):
        start = curat.find('{')
        end = curat.rfind('}')
        if start != -1 and end != -1:
            curat = curat[start:end + 1]
    return json.loads(curat)


FORMATE_ACCEPTATE = [
    'image/png',
    'image/jpeg',
    'image/jpg',
    'image/webp',
    'application/pdf',
]


def proceseaza_imagine_bon(fisier):
    nume = (getattr(fisier, 'name', '') or '').lower()

    try:
        fisier.seek(0)
    except Exception:
        pass

    if nume.endswith('.pdf'):
        pdf_bytes = fisier.read()
        if not pdf_bytes:
            raise ValueError('Fișierul PDF este gol.')

        try:
            with fitz.open(stream=pdf_bytes, filetype='pdf') as doc:
                if doc.page_count == 0:
                    raise ValueError('PDF-ul nu conține pagini valide.')
                pagina = doc[0]
                mat = fitz.Matrix(2, 2)
                pix = pagina.get_pixmap(matrix=mat)
                img_bytes = pix.tobytes('png')
            imagine = PIL.Image.open(io.BytesIO(img_bytes))
        except ValueError:
            raise
        except Exception as exc:
            raise ValueError('Fișierul PDF nu poate fi procesat.') from exc
    else:
        try:
            imagine = PIL.Image.open(fisier)
        except UnidentifiedImageError as exc:
            raise ValueError('Fișierul încărcat nu este o imagine validă.') from exc

    imagine = ImageOps.exif_transpose(imagine)
    if imagine.mode not in ['RGB', 'L']:
        imagine = imagine.convert('RGB')
    imagine.load()

    return imagine


def pregateste_bon_upload(uploaded_file):
    extensie = Path(uploaded_file.name).suffix.lower()
    if extensie not in ['.jpg', '.jpeg', '.png', '.webp', '.pdf']:
        raise ValueError('Fișierul trebuie să fie PNG, JPG, WebP sau PDF')

    content_type = (uploaded_file.content_type or '').lower()
    if content_type and content_type not in FORMATE_ACCEPTATE:
        raise ValueError('Fișierul trebuie să fie PNG, JPG, WebP sau PDF')

    file_bytes = uploaded_file.read()
    if not file_bytes:
        raise ValueError('Fișierul încărcat este gol.')

    fisier_pentru_procesare = BytesIO(file_bytes)
    fisier_pentru_procesare.name = uploaded_file.name
    proceseaza_imagine_bon(fisier_pentru_procesare)

    identificator = uuid.uuid4().hex
    fisier_original_relativ = f'bonuri/{identificator}{extensie or ".bin"}'
    fisier_original_salvat = default_storage.save(fisier_original_relativ, ContentFile(file_bytes))
    original_url = default_storage.url(fisier_original_salvat)
    original_path = default_storage.path(fisier_original_salvat)

    return {
        'original_url': original_url,
        'preview_url': original_url,
        'stored_path': fisier_original_salvat,
        'imagine_path': original_path,
        'filename': uploaded_file.name,
        'is_pdf': extensie == '.pdf',
    }


def analizeaza_bon_cu_gemini(imagine_path):
    prompt = (
        "Analizează acest bon de cumpărături și extrage:\n"
        "1. Numele magazinului\n"
        "2. Data cumpărăturii\n"
        "3. Lista produselor cu prețurile lor\n"
        "4. Totalul bonului\n"
        "5. Categoria sugerată (Food & Groceries, Electronics, etc.)\n"
        "Răspunde DOAR în format JSON cu structura:\n"
        "{\n"
        '  "magazin": "nume",\n'
        '  "data": "YYYY-MM-DD",\n'
        '  "produse": [{"nume": "...", "pret": 0.00}],\n'
        '  "total": 0.00,\n'
        '  "categorie_sugerata": "..."\n'
        "}"
    )

    genai.configure(api_key=settings.GEMINI_API_KEY)
    model = genai.GenerativeModel('gemini-flash-lite-latest')
    with open(imagine_path, 'rb') as fisier:
        imagine = proceseaza_imagine_bon(fisier)
    response = model.generate_content([prompt, imagine])
    response_text = (response.text or '').strip()
    return curata_json_text(response_text), response_text


def obtine_date_bon_din_sesiune(request):
    return request.session.get('bon_scan_data')


def salveaza_date_bon_in_sesiune(request, scan_data):
    request.session['bon_scan_data'] = scan_data
    request.session.modified = True


def sterge_date_bon_din_sesiune(request):
    if 'bon_scan_data' in request.session:
        del request.session['bon_scan_data']
        request.session.modified = True


@login_required
@check_limit('scanare')
def scaneaza_bon(request):
    categorii = Categorie.objects.all().order_by('nume')
    bon_scan_data = obtine_date_bon_din_sesiune(request)

    if request.method == 'POST':
        actiune = request.POST.get('action')

        if actiune == 'analizeaza':
            fisier = request.FILES.get('bon_fisier')
            if not fisier:
                messages.error(request, 'Alege un fișier PNG, JPG, WebP sau PDF înainte de analiză.')
                return redirect('scaneaza_bon')

            try:
                pregatit = pregateste_bon_upload(fisier)
                analiza_json, text_brut = analizeaza_bon_cu_gemini(pregatit['imagine_path'])

                magazin = str(analiza_json.get('magazin', '')).strip() or 'Bon scanat'
                data_text = str(analiza_json.get('data', '')).strip()
                produse = analiza_json.get('produse', []) or []
                total = Decimal(str(analiza_json.get('total', 0))).quantize(Decimal('0.01'))
                categorie_sugerata = str(analiza_json.get('categorie_sugerata', '')).strip()
                data_cumpararii = data_text or timezone.now().date().isoformat()

                categorie_initiala = Categorie.objects.filter(nume__iexact=categorie_sugerata).first()
                descriere = construieste_descriere_bon(produse, magazin, data_cumpararii, categorie_sugerata)

                scan_data = {
                    'file_url': pregatit['preview_url'],
                    'original_url': pregatit['original_url'],
                    'stored_path': pregatit['stored_path'],
                    'filename': pregatit['filename'],
                    'is_pdf': pregatit['is_pdf'],
                    'magazin': magazin,
                    'data_cumpararii': data_cumpararii,
                    'produse': produse,
                    'total': float(total),
                    'categorie_sugerata': categorie_sugerata,
                    'categorie_initiala_id': categorie_initiala.pk if categorie_initiala else None,
                    'titlu_propus': f'Bon - {magazin}',
                    'descriere_propusa': descriere,
                    'raw_response': text_brut,
                }
                salveaza_date_bon_in_sesiune(request, scan_data)
                ScanareLog.objects.create(utilizator=request.user)
                messages.success(request, 'Bonul a fost analizat cu succes.')
                return redirect('scaneaza_bon')
            except Exception as exc:
                messages.error(request, f'Nu s-a putut analiza bonul: {exc}')
                return redirect('scaneaza_bon')

        if actiune == 'salveaza':
            if not bon_scan_data:
                messages.error(request, 'Nu există date de bon salvate pentru confirmare.')
                return redirect('scaneaza_bon')

            titlu = request.POST.get('titlu', '').strip() or bon_scan_data.get('titlu_propus', 'Bon scanat')
            total_text = request.POST.get('total', '').strip()
            categorie_sugerata_text = request.POST.get('categorie_sugerata', '').strip()
            categorie_id = request.POST.get('categorie') or ''
            categorie = None
            if categorie_id:
                categorie = Categorie.objects.filter(pk=categorie_id).first()
            if not categorie and categorie_sugerata_text:
                categorie = Categorie.objects.filter(nume__iexact=categorie_sugerata_text).first()
            if not categorie and bon_scan_data.get('categorie_initiala_id'):
                categorie = Categorie.objects.filter(pk=bon_scan_data['categorie_initiala_id']).first()

            data_cumpararii_text = request.POST.get('data_cumpararii', bon_scan_data.get('data_cumpararii'))
            try:
                data_cumpararii = date.fromisoformat(data_cumpararii_text)
            except Exception:
                data_cumpararii = timezone.now().date()

            descriere = request.POST.get('descriere', '').strip() or bon_scan_data.get('descriere_propusa', '')
            try:
                suma = Decimal(total_text or str(bon_scan_data.get('total', 0))).quantize(Decimal('0.01'))
            except Exception:
                suma = Decimal(str(bon_scan_data.get('total', 0))).quantize(Decimal('0.01'))

            cheltuiala = Cheltuiala.objects.create(
                utilizator=request.user,
                categorie=categorie,
                titlu=titlu,
                suma=suma,
                data=data_cumpararii,
                descriere=descriere,
            )

            produse_bon = bon_scan_data.get('produse', []) or []
            produse_bon = produse_bon if isinstance(produse_bon, list) else []
            nr_produse = len([item for item in produse_bon if isinstance(item, dict)])
            pret_mediu_produs = rotunjeste_bani(suma / Decimal(nr_produse)) if nr_produse > 0 else rotunjeste_bani(suma)
            magazin_bon = str(bon_scan_data.get('magazin') or titlu or 'Bon scanat').strip()

            ReceiptInsight.objects.create(
                utilizator=request.user,
                magazin=magazin_bon[:160],
                data_bon=data_cumpararii,
                total=rotunjeste_bani(suma),
                nr_produse=nr_produse,
                pret_mediu_produs=pret_mediu_produs,
                produse_json=produse_bon,
            )

            sterge_date_bon_din_sesiune(request)
            messages.success(request, f'Cheltuiala "{cheltuiala.titlu}" a fost creată din bon!')
            return redirect('lista_cheltuieli')

    context = {
        'categorii': categorii,
        'bon_scan': bon_scan_data,
    }
    return render(request, 'spndix/scaneaza_bon.html', context)


def genereaza_analiza_ai(request, luna, an, force=False):
    existing = AIAnaliza.objects.filter(utilizator=request.user, luna=luna, an=an).first()
    if existing and not force:
        return existing, False

    if not settings.GEMINI_API_KEY:
        raise RuntimeError('Lipsește cheia GEMINI_API_KEY din .env')

    _, date_bugete_text, _, _, _ = rezumat_cheltuieli_ai(request, luna, an)

    try:
        profil = request.user.userprofile
        tip_gospodarie = profil.get_tip_gospodarie_display()
        nr_persoane = profil.nr_persoane
        are_copii = "Da" if profil.are_copii else "Nu"
        venit_lunar = profil.venit_lunar or "Nespecificat"
        obiectiv = profil.get_obiectiv_display()
    except UserProfile.DoesNotExist:
        tip_gospodarie = "Nespecificat"
        nr_persoane = 1
        are_copii = "Nu"
        venit_lunar = "Nespecificat"
        obiectiv = "Nespecificat"

    cheltuieli_luna = Cheltuiala.objects.filter(
        utilizator=request.user,
        data__month=luna,
        data__year=an,
    ).select_related('categorie').order_by('-data', '-id')

    date_cheltuieli_detaliate = construieste_cheltuieli_detaliate(cheltuieli_luna)
    date_luna_precedenta = construieste_date_luna_precedenta(request.user, luna, an)

    prompt = construieste_prompt_analiza(
        luna=luna_display(luna),
        an=an,
        tip_gospodarie=tip_gospodarie,
        nr_persoane=nr_persoane,
        are_copii=are_copii,
        venit_lunar=venit_lunar,
        obiectiv=obiectiv,
        date_cheltuieli_detaliate=date_cheltuieli_detaliate,
        date_bugete=date_bugete_text,
        date_luna_precedenta=date_luna_precedenta,
    )

    genai.configure(api_key=settings.GEMINI_API_KEY)
    analysis_text = curata_markdown(genereaza_text_gemini(prompt))

    if existing:
        existing.continut_analiza = analysis_text
        existing.creat_la = timezone.now()
        existing.save(update_fields=['continut_analiza', 'creat_la'])
        return existing, True

    return AIAnaliza.objects.create(
        utilizator=request.user,
        luna=luna,
        an=an,
        continut_analiza=analysis_text,
    ), True


@login_required
def dashboard(request):
    azi = timezone.now()
    azi_data = timezone.localdate()
    luna_curenta = azi.month
    an_curent = azi.year

    auto_adauga_subscriptions_lunare(request.user, azi=azi_data)

    total_luna = Cheltuiala.objects.filter(
        utilizator=request.user,
        data__month=luna_curenta,
        data__year=an_curent
    ).aggregate(total=Sum('suma'))['total'] or 0

    total_general = Cheltuiala.objects.filter(
        utilizator=request.user
    ).aggregate(total=Sum('suma'))['total'] or 0

    ultimele_cheltuieli = Cheltuiala.objects.filter(
        utilizator=request.user
    )[:5]

    bugete_curente = Budget.objects.filter(
        utilizator=request.user,
        luna=luna_curenta,
        an=an_curent,
    ).select_related('categorie')

    bugete_dashboard = [construieste_buget_dashboard(buget, request.user) for buget in bugete_curente]
    bugete_depasite = [buget for buget in bugete_dashboard if buget['excedat']]

    subscriptions_active = Subscription.objects.filter(utilizator=request.user, activ=True).select_related('categorie')
    total_subscriptions_lunar = sum(
        (suma_lunara_subscription(item) for item in subscriptions_active),
        Decimal('0'),
    )

    categorii_data = Cheltuiala.objects.filter(
        utilizator=request.user,
        data__month=luna_curenta,
        data__year=an_curent
    ).values('categorie__nume').annotate(total=Sum('suma'))

    pie_labels = [item['categorie__nume'] or 'Fără categorie' for item in categorii_data]
    pie_data = [float(item['total']) for item in categorii_data]

    bar_labels = []
    bar_data = []
    for i in range(5, -1, -1):
        if luna_curenta - i <= 0:
            luna = luna_curenta - i + 12
            an = an_curent - 1
        else:
            luna = luna_curenta - i
            an = an_curent

        total = Cheltuiala.objects.filter(
            utilizator=request.user,
            data__month=luna,
            data__year=an
        ).aggregate(total=Sum('suma'))['total'] or 0

        nume_luna = ['Ian', 'Feb', 'Mar', 'Apr', 'Mai', 'Iun',
                     'Iul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'][luna - 1]
        bar_labels.append(nume_luna)
        bar_data.append(float(total))

    forecast_alerts = calculate_forecasts(request.user)
    smart_actions = sincronizeaza_smart_actions_din_alerte(request.user)
    smart_actions_pending = [item for item in smart_actions if item.status == 'pending'][:4]
    onboarding_context = obtine_onboarding_context(request.user)
    receipt_intelligence = calculeaza_receipt_intelligence(request.user, azi=azi_data)
    household_summary = obtine_household_summary(request.user)

    goals_active = SavingsGoal.objects.filter(utilizator=request.user, activ=True)
    inceput_saptamana = timezone.localdate() - timedelta(days=6)
    contributii_saptamana = GoalContribution.objects.filter(
        goal__utilizator=request.user,
        goal__activ=True,
        data__gte=inceput_saptamana,
    ).values('goal_id').annotate(total=Sum('suma'))
    contributii_saptamana_map = {
        item['goal_id']: Decimal(item['total'] or 0)
        for item in contributii_saptamana
    }
    goals_top = [
        construieste_goal_status(goal, contributii_saptamana_map)
        for goal in goals_active.order_by('-creat_la')
    ]
    goals_top = sorted(goals_top, key=lambda item: item['procent_real'], reverse=True)[:3]
    total_economisit_goals = goals_active.aggregate(total=Sum('suma_curenta'))['total'] or Decimal('0')

    context = {
        'total_luna': total_luna,
        'total_general': total_general,
        'ultimele_cheltuieli': ultimele_cheltuieli,
        'bugete_dashboard': bugete_dashboard,
        'bugete_depasite': bugete_depasite,
        'pie_labels': json.dumps(pie_labels),
        'pie_data': json.dumps(pie_data),
        'bar_labels': json.dumps(bar_labels),
        'bar_data': json.dumps(bar_data),
        'luna_curenta': azi.strftime('%B %Y'),
        'forecast_alerts': forecast_alerts,
        'smart_actions_pending': smart_actions_pending,
        'smart_actions_total_pending': len([item for item in smart_actions if item.status == 'pending']),
        'onboarding': onboarding_context,
        'receipt_intelligence': receipt_intelligence,
        'household_summary': household_summary,
        'goals_top': goals_top,
        'total_economisit_goals': rotunjeste_bani(total_economisit_goals),
        'goals_active_count': goals_active.count(),
        'subscriptions_active_count': subscriptions_active.count(),
        'subscriptions_total_lunar': rotunjeste_bani(total_subscriptions_lunar),
    }
    return render(request, 'spndix/dashboard.html', context)


@login_required
def marca_citita(request, pk):
    alerta = get_object_or_404(ForecastAlert, pk=pk, utilizator=request.user)
    if not alerta.citita:
        alerta.citita = True
        alerta.save(update_fields=['citita'])

    next_url = request.POST.get('next') or request.GET.get('next')
    if next_url:
        return redirect(next_url)
    return redirect('dashboard')


@login_required
def upgrade(request):
    plan_curent = obtine_user_plan(request.user)
    return render(
        request,
        'spndix/upgrade.html',
        {
            'plan_curent': plan_curent,
        },
    )


@login_required
def checkout(request, plan):
    plan = (plan or '').lower()
    if plan == 'pro':
        suma = 1000
        nume_plan = 'Spndix Pro'
    elif plan == 'family':
        suma = 2000
        nume_plan = 'Spndix Family'
    else:
        messages.error(request, 'Plan invalid selectat.')
        return redirect('upgrade')

    if not settings.STRIPE_SECRET_KEY or not settings.STRIPE_PUBLIC_KEY:
        messages.error(request, 'Stripe nu este configurat complet. Verifică variabilele din .env.')
        return redirect('upgrade')

    stripe.api_key = settings.STRIPE_SECRET_KEY

    checkout_kwargs = {
        'payment_method_types': ['card'],
        'line_items': [{
            'price_data': {
                'currency': 'ron',
                'product_data': {
                    'name': nume_plan,
                },
                'unit_amount': suma,
                'recurring': {
                    'interval': 'month',
                },
            },
            'quantity': 1,
        }],
        'mode': 'subscription',
        'success_url': request.build_absolute_uri('/upgrade/success/?session_id={CHECKOUT_SESSION_ID}'),
        'cancel_url': request.build_absolute_uri('/upgrade/'),
        'metadata': {
            'user_id': str(request.user.id),
            'plan': plan,
        },
    }

    if request.user.email:
        checkout_kwargs['customer_email'] = request.user.email

    try:
        session = stripe.checkout.Session.create(**checkout_kwargs)
    except Exception as exc:
        messages.error(request, f'Nu s-a putut crea sesiunea Stripe: {exc}')
        return redirect('upgrade')

    return redirect(session.url)


@login_required
def upgrade_success(request):
    session_id = request.GET.get('session_id')
    if not session_id:
        messages.error(request, 'Lipsește sesiunea Stripe pentru confirmarea plății.')
        return redirect('upgrade')

    if not settings.STRIPE_SECRET_KEY:
        messages.error(request, 'Stripe nu este configurat complet. Verifică variabilele din .env.')
        return redirect('upgrade')

    stripe.api_key = settings.STRIPE_SECRET_KEY
    try:
        session = stripe.checkout.Session.retrieve(session_id)
    except Exception as exc:
        messages.error(request, f'Nu s-a putut valida sesiunea Stripe: {exc}')
        return redirect('upgrade')

    if session.get('status') != 'complete':
        messages.warning(request, 'Plata nu este încă finalizată.')
        return redirect('upgrade')

    plan_selectat = (session.get('metadata', {}) or {}).get('plan', 'pro').lower()
    if plan_selectat not in ['pro', 'family']:
        plan_selectat = 'pro'

    user_plan = obtine_user_plan(request.user)
    user_plan.plan = plan_selectat
    user_plan.activ = True
    user_plan.data_expirare = timezone.localdate() + timedelta(days=30)
    user_plan.stripe_customer_id = session.get('customer') or user_plan.stripe_customer_id
    subscription_id = session.get('subscription')
    if isinstance(subscription_id, str):
        user_plan.stripe_subscription_id = subscription_id
    user_plan.save(
        update_fields=[
            'plan',
            'activ',
            'data_expirare',
            'stripe_customer_id',
            'stripe_subscription_id',
        ]
    )

    return render(
        request,
        'spndix/upgrade_success.html',
        {
            'plan_selectat': plan_selectat,
        },
    )


@csrf_exempt
def stripe_webhook(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    if not settings.STRIPE_SECRET_KEY or not settings.STRIPE_WEBHOOK_SECRET:
        return JsonResponse({'error': 'Stripe webhook is not configured'}, status=500)

    stripe.api_key = settings.STRIPE_SECRET_KEY
    payload = request.body
    signature = request.META.get('HTTP_STRIPE_SIGNATURE', '')

    try:
        event = stripe.Webhook.construct_event(payload, signature, settings.STRIPE_WEBHOOK_SECRET)
    except ValueError:
        return JsonResponse({'error': 'Invalid payload'}, status=400)
    except stripe.error.SignatureVerificationError:
        return JsonResponse({'error': 'Invalid signature'}, status=400)

    event_type = event.get('type')
    data_object = (event.get('data') or {}).get('object', {})

    if event_type == 'checkout.session.completed':
        metadata = data_object.get('metadata') or {}
        user_id = metadata.get('user_id')
        plan_selectat = str(metadata.get('plan', 'pro')).lower()
        if plan_selectat not in ['pro', 'family']:
            plan_selectat = 'pro'

        if user_id:
            user = User.objects.filter(pk=user_id).first()
            if user:
                user_plan = obtine_user_plan(user)
                user_plan.plan = plan_selectat
                user_plan.activ = True
                user_plan.data_expirare = timezone.localdate() + timedelta(days=30)
                user_plan.stripe_customer_id = data_object.get('customer') or user_plan.stripe_customer_id
                subscription_id = data_object.get('subscription')
                if isinstance(subscription_id, str):
                    user_plan.stripe_subscription_id = subscription_id
                user_plan.save(
                    update_fields=[
                        'plan',
                        'activ',
                        'data_expirare',
                        'stripe_customer_id',
                        'stripe_subscription_id',
                    ]
                )

    if event_type in ['customer.subscription.deleted', 'customer.subscription.updated']:
        subscription_id = data_object.get('id')
        subscription_status = data_object.get('status')
        if subscription_id:
            user_plan = UserPlan.objects.filter(stripe_subscription_id=subscription_id).first()
            if user_plan and subscription_status in ['canceled', 'unpaid', 'incomplete_expired']:
                user_plan.plan = 'free'
                user_plan.activ = True
                user_plan.data_expirare = None
                user_plan.save(update_fields=['plan', 'activ', 'data_expirare'])

    return JsonResponse({'status': 'ok'})


@login_required
def lista_smart_actions(request):
    calculate_forecasts(request.user)
    obtine_onboarding_context(request.user)
    actions = sincronizeaza_smart_actions_din_alerte(request.user)

    context = {
        'actions_pending': [item for item in actions if item.status == 'pending'],
        'actions_done': [item for item in actions if item.status == 'done'],
        'actions_dismissed': [item for item in actions if item.status == 'dismissed'],
        'impact_total_pending': rotunjeste_bani(
            sum((Decimal(item.impact_estimat or 0) for item in actions if item.status == 'pending'), Decimal('0'))
        ),
    }
    return render(request, 'spndix/smart_actions/lista.html', context)


@login_required
def actualizeaza_smart_action(request, pk, actiune):
    smart_action = get_object_or_404(SmartAction, pk=pk, utilizator=request.user)

    if request.method == 'POST':
        if actiune == 'done':
            smart_action.status = 'done'
            smart_action.completata_la = timezone.now()
            smart_action.save(update_fields=['status', 'completata_la'])
            messages.success(request, 'Smart action marcat ca finalizat.')
        elif actiune == 'dismiss':
            smart_action.status = 'dismissed'
            smart_action.save(update_fields=['status'])
            messages.info(request, 'Smart action marcat ca ignorat.')
        elif actiune == 'reopen':
            smart_action.status = 'pending'
            smart_action.completata_la = None
            smart_action.save(update_fields=['status', 'completata_la'])
            messages.success(request, 'Smart action reactivat.')

    next_url = request.POST.get('next') or request.GET.get('next')
    if next_url:
        return redirect(next_url)
    return redirect('lista_smart_actions')


@login_required
def onboarding(request):
    onboarding_context = obtine_onboarding_context(request.user)
    journey = onboarding_context['journey']

    if request.method == 'POST':
        actiune = request.POST.get('action')
        if actiune == 'ascunde':
            journey.ascuns = True
            journey.save(update_fields=['ascuns'])
            messages.info(request, 'Onboarding ascuns. Îl poți reactiva oricând.')
            return redirect('dashboard')
        if actiune == 'reactiveaza':
            journey.ascuns = False
            journey.save(update_fields=['ascuns'])
            messages.success(request, 'Onboarding reactivat.')
            return redirect('onboarding')

    onboarding_context = obtine_onboarding_context(request.user)
    return render(request, 'spndix/onboarding.html', {'onboarding': onboarding_context})


@login_required
def household(request):
    membership = HouseholdMember.objects.filter(
        utilizator=request.user,
        activ=True,
    ).select_related('household').first()
    household_curenta = membership.household if membership else None

    create_form = HouseholdCreateForm(prefix='create')
    add_member_form = HouseholdAddMemberForm(prefix='member')

    if request.method == 'POST':
        actiune = request.POST.get('action')

        if actiune == 'create':
            create_form = HouseholdCreateForm(request.POST, prefix='create')
            if create_form.is_valid():
                nume = create_form.cleaned_data['nume'].strip()
                household_nou = Household.objects.create(
                    nume=nume,
                    owner=request.user,
                )
                HouseholdMember.objects.create(
                    household=household_nou,
                    utilizator=request.user,
                    rol='owner',
                    responsabilitate='Coordonare buget',
                )
                messages.success(request, f"Gospodăria '{nume}' a fost creată.")
                return redirect('household')

        if actiune == 'add_member':
            if not household_curenta or household_curenta.owner_id != request.user.id:
                messages.error(request, 'Doar owner-ul gospodăriei poate adăuga membri.')
                return redirect('household')

            add_member_form = HouseholdAddMemberForm(request.POST, prefix='member')
            if add_member_form.is_valid():
                username = add_member_form.cleaned_data['username']
                rol = add_member_form.cleaned_data['rol']
                responsabilitate = add_member_form.cleaned_data['responsabilitate']
                user_nou = User.objects.filter(username=username).first()

                if not user_nou:
                    messages.error(request, 'Utilizatorul nu există.')
                    return redirect('household')

                HouseholdMember.objects.update_or_create(
                    household=household_curenta,
                    utilizator=user_nou,
                    defaults={
                        'rol': rol,
                        'responsabilitate': responsabilitate,
                        'activ': True,
                    },
                )
                messages.success(request, f"Utilizatorul {username} a fost adăugat în gospodărie.")
                return redirect('household')

        if actiune == 'remove_member' and household_curenta and household_curenta.owner_id == request.user.id:
            member_id = request.POST.get('member_id')
            membru = HouseholdMember.objects.filter(pk=member_id, household=household_curenta, activ=True).first()
            if membru:
                if membru.utilizator_id == request.user.id:
                    messages.error(request, 'Owner-ul nu poate fi eliminat din propria gospodărie.')
                else:
                    membru.activ = False
                    membru.save(update_fields=['activ'])
                    messages.success(request, f"{membru.utilizator.username} a fost scos din gospodărie.")
            return redirect('household')

        if actiune == 'leave' and membership and membership.rol != 'owner':
            membership.activ = False
            membership.save(update_fields=['activ'])
            messages.success(request, 'Ai ieșit din gospodărie.')
            return redirect('household')

    membership = HouseholdMember.objects.filter(
        utilizator=request.user,
        activ=True,
    ).select_related('household').first()
    household_curenta = membership.household if membership else None

    membri_household = []
    cheltuieli_membri = []
    stats_household = None
    if household_curenta:
        membri_household = list(
            household_curenta.membri.filter(activ=True)
            .select_related('utilizator')
            .order_by('rol', 'utilizator__username')
        )
        member_ids = [item.utilizator_id for item in membri_household]
        azi = timezone.localdate()
        cheltuieli_membri = list(
            Cheltuiala.objects.filter(
                utilizator_id__in=member_ids,
                data__month=azi.month,
                data__year=azi.year,
            )
            .values('utilizator__username')
            .annotate(total=Sum('suma'))
            .order_by('-total')
        )
        total_luna = sum((Decimal(item['total'] or 0) for item in cheltuieli_membri), Decimal('0'))
        stats_household = {
            'membri_count': len(membri_household),
            'total_luna': rotunjeste_bani(total_luna),
        }

    context = {
        'membership': membership,
        'household_curenta': household_curenta,
        'membri_household': membri_household,
        'cheltuieli_membri': cheltuieli_membri,
        'stats_household': stats_household,
        'create_form': create_form,
        'add_member_form': add_member_form,
        'este_owner': bool(household_curenta and household_curenta.owner_id == request.user.id),
    }
    return render(request, 'spndix/household.html', context)


@login_required
def lista_goals(request):
    goals = SavingsGoal.objects.filter(utilizator=request.user).order_by('-activ', 'data_tinta', '-creat_la')
    inceput_saptamana = timezone.localdate() - timedelta(days=6)
    contributii_saptamana = GoalContribution.objects.filter(
        goal__utilizator=request.user,
        data__gte=inceput_saptamana,
    ).values('goal_id').annotate(total=Sum('suma'))
    contributii_saptamana_map = {
        item['goal_id']: Decimal(item['total'] or 0)
        for item in contributii_saptamana
    }

    goal_cards = [construieste_goal_status(goal, contributii_saptamana_map) for goal in goals]
    total_economisit = SavingsGoal.objects.filter(
        utilizator=request.user,
        activ=True,
    ).aggregate(total=Sum('suma_curenta'))['total'] or Decimal('0')

    context = {
        'goal_cards': goal_cards,
        'total_economisit': rotunjeste_bani(total_economisit),
    }
    return render(request, 'spndix/goals/lista.html', context)


@login_required
def adauga_goal(request):
    if request.method == 'POST':
        form = SavingsGoalForm(request.POST)
        if form.is_valid():
            goal = form.save(commit=False)
            goal.utilizator = request.user
            goal.save()
            messages.success(request, f"Goal-ul '{goal.titlu}' a fost creat.")
            return redirect('lista_goals')
    else:
        form = SavingsGoalForm()

    return render(request, 'spndix/form.html', {'form': form, 'titlu': 'Adaugă Goal', 'cancel_url': 'lista_goals'})


@login_required
def editeaza_goal(request, pk):
    goal = get_object_or_404(SavingsGoal, pk=pk, utilizator=request.user)
    if request.method == 'POST':
        form = SavingsGoalForm(request.POST, instance=goal)
        if form.is_valid():
            goal = form.save()
            messages.success(request, f"Goal-ul '{goal.titlu}' a fost actualizat.")
            return redirect('lista_goals')
    else:
        form = SavingsGoalForm(instance=goal)

    return render(request, 'spndix/form.html', {'form': form, 'titlu': 'Editează Goal', 'cancel_url': 'lista_goals'})


@login_required
def sterge_goal(request, pk):
    goal = get_object_or_404(SavingsGoal, pk=pk, utilizator=request.user)
    if request.method == 'POST':
        titlu = goal.titlu
        goal.delete()
        messages.success(request, f"Goal-ul '{titlu}' a fost șters.")
        return redirect('lista_goals')

    return render(request, 'spndix/goals/confirmare_stergere.html', {'goal': goal})


@login_required
def adauga_contributie(request, pk):
    goal = get_object_or_404(SavingsGoal, pk=pk, utilizator=request.user)
    contributii = goal.contributii.all()[:15]
    goal_status = construieste_goal_status(goal)

    if request.method == 'POST':
        form = GoalContributionForm(request.POST)
        if form.is_valid():
            procent_inainte = procent_goal(goal.suma_curenta, goal.suma_tinta)

            contributie = form.save(commit=False)
            contributie.goal = goal
            contributie.data = timezone.localdate()
            contributie.save()

            goal.suma_curenta = rotunjeste_bani(Decimal(goal.suma_curenta or 0) + Decimal(contributie.suma or 0))
            goal.save(update_fields=['suma_curenta'])

            procent_dupa = procent_goal(goal.suma_curenta, goal.suma_tinta)
            notifica_milestone_goal(goal, procent_inainte, procent_dupa)

            messages.success(request, f"Ai adăugat {rotunjeste_bani(contributie.suma)} RON la goal-ul '{goal.titlu}'.")
            return redirect('lista_goals')
    else:
        form = GoalContributionForm()

    context = {
        'goal': goal,
        'goal_status': goal_status,
        'form': form,
        'contributii': contributii,
    }
    return render(request, 'spndix/goals/adauga_contributie.html', context)


def construieste_subscription_card(subscription):
    return {
        'subscription': subscription,
        'suma_lunara': suma_lunara_subscription(subscription),
        'urmatoarea_plata': subscription.urmatoarea_plata,
    }


@login_required
def lista_subscriptions(request):
    azi = timezone.localdate()
    sincronizeaza_urmatoare_plati_subscriptions(request.user, referinta=azi)

    if request.method == 'POST':
        form = SubscriptionForm(request.POST)
        if form.is_valid():
            subscription = form.save(commit=False)
            subscription.utilizator = request.user
            subscription.frecventa = 'lunar'
            subscription.activ = True
            subscription.detectat_automat = False
            if not subscription.ziua_lunii:
                subscription.ziua_lunii = 1
            subscription.urmatoarea_plata = urmatoarea_data_subscription(subscription, referinta=azi)
            subscription.save()
            messages.success(request, f"Abonamentul '{subscription.nume}' a fost adăugat.")
            return redirect('lista_subscriptions')
    else:
        form = SubscriptionForm(initial={'ziua_lunii': 1})

    subscriptions = list(
        Subscription.objects.filter(utilizator=request.user)
        .select_related('categorie')
        .order_by('-activ', 'urmatoarea_plata', 'nume')
    )
    subscriptions_active = [item for item in subscriptions if item.activ]
    subscriptions_inactive = [item for item in subscriptions if not item.activ]
    total_lunar = sum((suma_lunara_subscription(item) for item in subscriptions_active), Decimal('0'))

    venit_lunar = None
    try:
        profil = request.user.userprofile
        if profil.venit_lunar:
            venit_lunar = Decimal(profil.venit_lunar)
    except UserProfile.DoesNotExist:
        venit_lunar = None

    prag_20_venit = None
    depaseste_prag_venit = False
    if venit_lunar and venit_lunar > 0:
        prag_20_venit = rotunjeste_bani(venit_lunar * Decimal('0.20'))
        depaseste_prag_venit = total_lunar > prag_20_venit

    context = {
        'form': form,
        'active_subscription_cards': [construieste_subscription_card(item) for item in subscriptions_active],
        'inactive_subscription_cards': [construieste_subscription_card(item) for item in subscriptions_inactive],
        'subscriptions_active_count': len(subscriptions_active),
        'subscriptions_total_lunar': rotunjeste_bani(total_lunar),
        'venit_lunar': rotunjeste_bani(venit_lunar) if venit_lunar else None,
        'prag_20_venit': prag_20_venit,
        'depaseste_prag_venit': depaseste_prag_venit,
    }
    return render(request, 'spndix/subscriptions/lista.html', context)


@login_required
def editeaza_subscription(request, pk):
    subscription = get_object_or_404(Subscription, pk=pk, utilizator=request.user)

    if request.method == 'POST':
        form = SubscriptionForm(request.POST, instance=subscription)
        if form.is_valid():
            subscription = form.save(commit=False)
            subscription.frecventa = 'lunar'
            if not subscription.ziua_lunii:
                subscription.ziua_lunii = 1
            if subscription.activ:
                subscription.urmatoarea_plata = urmatoarea_data_subscription(subscription, referinta=timezone.localdate())
            subscription.save()
            messages.success(request, f"Abonamentul '{subscription.nume}' a fost actualizat.")
            return redirect('lista_subscriptions')
    else:
        form = SubscriptionForm(instance=subscription)

    context = {
        'form': form,
        'titlu': 'Editează abonament',
        'cancel_url': 'lista_subscriptions',
    }
    return render(request, 'spndix/form.html', context)


@login_required
def anuleaza_subscription(request, pk):
    subscription = get_object_or_404(Subscription, pk=pk, utilizator=request.user)

    if request.method == 'POST' and subscription.activ:
        subscription.activ = False
        subscription.urmatoarea_plata = None
        subscription.save(update_fields=['activ', 'urmatoarea_plata'])
        messages.success(request, f"Monitorizarea abonamentului '{subscription.nume}' a fost anulată.")

    return redirect('lista_subscriptions')


@login_required
def activeaza_subscription(request, pk):
    subscription = get_object_or_404(Subscription, pk=pk, utilizator=request.user)

    if request.method == 'POST' and not subscription.activ:
        subscription.activ = True
        if not subscription.ziua_lunii:
            subscription.ziua_lunii = 1
        subscription.urmatoarea_plata = urmatoarea_data_subscription(subscription, referinta=timezone.localdate())
        subscription.save(update_fields=['activ', 'ziua_lunii', 'urmatoarea_plata'])
        messages.success(request, f"Monitorizarea abonamentului '{subscription.nume}' a fost reactivată.")

    return redirect('lista_subscriptions')


@login_required
def sterge_subscription(request, pk):
    subscription = get_object_or_404(Subscription, pk=pk, utilizator=request.user)

    if request.method == 'POST':
        nume = subscription.nume
        subscription.delete()
        messages.success(request, f"Abonamentul '{nume}' a fost șters.")

    return redirect('lista_subscriptions')

@login_required
def lista_cheltuieli(request):
    cheltuieli = Cheltuiala.objects.filter(utilizator=request.user).select_related('categorie').order_by('-data', '-id')
    luna_selectata, an_selectat = obtine_filtre_luna_an(request)
    categorie_selectata = request.GET.get('categorie')

    cheltuieli_filtrate = cheltuieli.filter(data__month=luna_selectata, data__year=an_selectat)
    if categorie_selectata:
        try:
            categorie_selectata = int(categorie_selectata)
            cheltuieli_filtrate = cheltuieli_filtrate.filter(categorie_id=categorie_selectata)
        except (TypeError, ValueError):
            categorie_selectata = None

    total_cheltuit = cheltuieli_filtrate.aggregate(total=Sum('suma'))['total'] or 0
    ani_disponibili = sorted(
        set(
            Cheltuiala.objects.filter(utilizator=request.user).dates('data', 'year')
        ),
        reverse=True,
    )
    ani_disponibili = [data.year for data in ani_disponibili]
    if an_selectat not in ani_disponibili:
        ani_disponibili.insert(0, an_selectat)

    context = {
        'cheltuieli': cheltuieli_filtrate,
        'total_cheltuit': total_cheltuit,
        'luna_selectata': luna_selectata,
        'an_selectat': an_selectat,
        'categorie_selectata': categorie_selectata,
        'luni': LUNA_CHOICES,
        'ani_disponibili': sorted(set(ani_disponibili), reverse=True),
    }
    return render(request, 'spndix/lista.html', context)


@login_required
@check_limit('analiza')
def analiza_ai(request):
    luna_selectata, an_selectat = obtine_filtre_luna_an(request)
    cheltuieli_count = Cheltuiala.objects.filter(
        utilizator=request.user,
        data__month=luna_selectata,
        data__year=an_selectat,
    ).count()
    warning_date_insuficiente = cheltuieli_count < 10

    if request.method == 'POST':
        luna_selectata = int(request.POST.get('month', luna_selectata))
        an_selectat = int(request.POST.get('year', an_selectat))
        force = request.POST.get('force') == '1'

        try:
            analiza, created = genereaza_analiza_ai(request, luna_selectata, an_selectat, force=force)
            if created and force:
                messages.success(request, 'Analiza AI a fost regenerată cu succes.')
            elif created:
                messages.success(request, 'Analiza AI a fost generată cu succes.')
            else:
                messages.info(request, 'Analiza existentă a fost încărcată.')
            return redirect(f"{request.path}?month={luna_selectata}&year={an_selectat}")
        except Exception as exc:
            messages.error(request, f'Nu s-a putut genera analiza AI: {exc}')
            return redirect(f"{request.path}?month={luna_selectata}&year={an_selectat}")

    analiza = AIAnaliza.objects.filter(
        utilizator=request.user,
        luna=luna_selectata,
        an=an_selectat,
    ).first()

    if analiza:
        context = {
            'luna_selectata': luna_selectata,
            'an_selectat': an_selectat,
            'ani_disponibili': luna_ani_disponibili(request),
            'luni': LUNA_CHOICES,
            'analiza': analiza,
            'analiza_exista': True,
            'creata_azi': False,
            'cheltuieli_count': cheltuieli_count,
            'warning_date_insuficiente': warning_date_insuficiente,
        }
    else:
        date_cheltuieli_text, date_bugete_text, cheltuieli, totaluri_categorii, bugete_detaliate = rezumat_cheltuieli_ai(request, luna_selectata, an_selectat)
        context = {
            'luna_selectata': luna_selectata,
            'an_selectat': an_selectat,
            'ani_disponibili': luna_ani_disponibili(request),
            'luni': LUNA_CHOICES,
            'analiza': None,
            'analiza_exista': False,
            'creata_azi': False,
            'date_cheltuieli_text': date_cheltuieli_text,
            'date_bugete_text': date_bugete_text,
            'cheltuieli_count': cheltuieli.count(),
            'categorii_count': len(totaluri_categorii),
            'bugete_count': len(bugete_detaliate),
            'warning_date_insuficiente': warning_date_insuficiente,
        }

    return render(request, 'spndix/analiza.html', context)


@login_required
@check_limit('export')
def export_cheltuieli_excel(request):
    cheltuieli, luna, an = cheltuieli_din_filtre(request)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Cheltuieli'

    ws.merge_cells('A1:E1')
    ws['A1'] = 'Spndix - Export Cheltuieli'
    ws['A1'].font = Font(bold=True, color='FFFFFF', size=14)
    ws['A1'].fill = PatternFill('solid', fgColor='1F4E78')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.append([f'Perioada: {luna_display(luna)} {an}', '', '', '', ''])
    ws.append(['Titlu', 'Categorie', 'Suma', 'Data', 'Descriere'])

    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='D9D9D9')
    total = 0

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_index = 4
    for cheltuiala in cheltuieli:
        categorie = cheltuiala.categorie.nume if cheltuiala.categorie else 'Fără categorie'
        ws.append([
            cheltuiala.titlu,
            categorie,
            float(cheltuiala.suma),
            cheltuiala.data.strftime('%d.%m.%Y'),
            cheltuiala.descriere or '',
        ])
        total += float(cheltuiala.suma)
        for col in range(1, 6):
            ws.cell(row=row_index, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row_index += 1

    ws.append(['', '', 'Total', total, ''])
    total_row = ws.max_row
    for cell in ws[total_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='D9EAF7')
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for column, width in {'A': 28, 'B': 26, 'C': 14, 'D': 14, 'E': 36}.items():
        ws.column_dimensions[column].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nume_fisier_export(luna, an, "xlsx")}"'

    output = BytesIO()
    wb.save(output)
    response.write(output.getvalue())
    ExportLog.objects.create(utilizator=request.user, tip='excel')
    return response


@login_required
@check_limit('export')
def export_cheltuieli_pdf(request):
    cheltuieli, luna, an = cheltuieli_din_filtre(request)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.2 * cm,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='SpndixTitle', parent=styles['Title'], fontSize=18, leading=22, textColor=colors.HexColor('#1F4E78')))
    styles.add(ParagraphStyle(name='SpndixSub', parent=styles['Normal'], fontSize=10, leading=12, textColor=colors.HexColor('#555555')))

    elements = [
        Paragraph('Spndix', styles['SpndixTitle']),
        Paragraph(f'Export Cheltuieli - {luna_display(luna)} {an}', styles['SpndixSub']),
        Spacer(1, 0.4 * cm),
    ]

    data = [['Titlu', 'Categorie', 'Suma', 'Data', 'Descriere']]
    total = 0
    for cheltuiala in cheltuieli:
        categorie = cheltuiala.categorie.nume if cheltuiala.categorie else 'Fără categorie'
        data.append([
            cheltuiala.titlu,
            categorie,
            f"{cheltuiala.suma:.2f} RON",
            cheltuiala.data.strftime('%d.%m.%Y'),
            cheltuiala.descriere or '',
        ])
        total += float(cheltuiala.suma)

    data.append(['', '', f'Total: {total:.2f} RON', '', ''])

    table = Table(data, colWidths=[6.0 * cm, 5.0 * cm, 3.5 * cm, 3.2 * cm, 7.5 * cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.whitesmoke, colors.HexColor('#EAF2F8')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9EAF7')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('SPAN', (0, -1), (1, -1)),
        ('ALIGN', (0, -1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))

    elements.append(table)
    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nume_fisier_export(luna, an, "pdf")}"'
    response.write(pdf)
    ExportLog.objects.create(utilizator=request.user, tip='pdf')
    return response

@login_required
def adauga_cheltuiala(request):
    if request.method == 'POST':
        form = CheltuialaForm(request.POST)
        if form.is_valid():
            cheltuiala = form.save(commit=False)
            cheltuiala.utilizator = request.user
            cheltuiala.save()
            messages.success(request, f'Cheltuiala "{cheltuiala.titlu}" a fost adăugată!')
            return redirect('lista_cheltuieli')
    else:
        form = CheltuialaForm()
    return render(
        request,
        'spndix/form.html',
        {
            'form': form,
            'titlu': 'Adaugă Cheltuială',
            'cancel_url': 'lista_cheltuieli',
            'show_scan_receipt_cta': True,
        },
    )

@login_required
def editeaza_cheltuiala(request, pk):
    cheltuiala = get_object_or_404(Cheltuiala, pk=pk, utilizator=request.user)
    if request.method == 'POST':
        form = CheltuialaForm(request.POST, instance=cheltuiala)
        if form.is_valid():
            form.save()
            messages.success(request, f'Cheltuiala "{cheltuiala.titlu}" a fost actualizată!')
            return redirect('lista_cheltuieli')
    else:
        form = CheltuialaForm(instance=cheltuiala)
    return render(request, 'spndix/form.html', {'form': form, 'titlu': 'Editează Cheltuială', 'cancel_url': 'lista_cheltuieli'})

@login_required
def sterge_cheltuiala(request, pk):
    cheltuiala = get_object_or_404(Cheltuiala, pk=pk, utilizator=request.user)
    if request.method == 'POST':
        cheltuiala.delete()
        messages.success(request, f'Cheltuiala "{cheltuiala.titlu}" a fost ștearsă!')
        return redirect('lista_cheltuieli')
    return render(request, 'spndix/confirmare_stergere.html', {'cheltuiala': cheltuiala, 'back_url': 'lista_cheltuieli'})


@login_required
def lista_bugete(request):
    bugete = Budget.objects.filter(utilizator=request.user).select_related('categorie')
    bugete_utile = [construieste_buget_dashboard(buget, request.user) for buget in bugete]
    return render(request, 'spndix/bugete/lista.html', {'bugete': bugete_utile})


@login_required
def adauga_buget(request):
    if request.method == 'POST':
        form = BudgetForm(request.POST)
        if form.is_valid():
            buget = form.save(commit=False)
            buget.utilizator = request.user
            buget.save()
            messages.success(request, f'Bugetul pentru {buget.categorie.nume} a fost adăugat!')
            return redirect('lista_bugete')
    else:
        form = BudgetForm()
    return render(request, 'spndix/form.html', {'form': form, 'titlu': 'Adaugă Buget', 'cancel_url': 'lista_bugete'})


@login_required
def editeaza_buget(request, pk):
    buget = get_object_or_404(Budget, pk=pk, utilizator=request.user)
    if request.method == 'POST':
        form = BudgetForm(request.POST, instance=buget)
        if form.is_valid():
            form.save()
            messages.success(request, f'Bugetul pentru {buget.categorie.nume} a fost actualizat!')
            return redirect('lista_bugete')
    else:
        form = BudgetForm(instance=buget)
    return render(request, 'spndix/form.html', {'form': form, 'titlu': 'Editează Buget', 'cancel_url': 'lista_bugete'})


@login_required
def sterge_buget(request, pk):
    buget = get_object_or_404(Budget, pk=pk, utilizator=request.user)
    if request.method == 'POST':
        buget.delete()
        messages.success(request, 'Bugetul a fost șters!')
        return redirect('lista_bugete')
    return render(request, 'spndix/confirmare_stergere.html', {'cheltuiala': buget, 'back_url': 'lista_bugete'})


@login_required
def profil(request):
    profil_user, _ = UserProfile.objects.get_or_create(utilizator=request.user)

    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=profil_user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profilul a fost salvat cu succes.')
            return redirect('profil')
    else:
        form = UserProfileForm(instance=profil_user)

    return render(request, 'spndix/profil.html', {'form': form})


def register(request):
    if request.user.is_authenticated:
        return redirect('lista_cheltuieli')
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, f'Cont creat cu succes! Bun venit, {user.username}!')
            return redirect('lista_cheltuieli')
    else:
        form = RegisterForm()
    return render(request, 'spndix/register.html', {'form': form})